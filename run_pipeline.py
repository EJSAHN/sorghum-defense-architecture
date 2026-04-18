#!/usr/bin/env python3
"""
Generate the supplementary data workbook for the sorghum defense architecture study.

The pipeline reads the three input data files, computes derived tables, and writes
a single Excel workbook. It does not generate publication figures.
"""

from __future__ import annotations

import argparse
import math
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from scipy.stats import chi2, fisher_exact, mannwhitneyu
from sklearn.decomposition import PCA
from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score
from statsmodels.miscmodels.ordinal_model import OrderedModel
from statsmodels.stats.multitest import multipletests

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT_ARCHITECTURE = "8leaf_final.xlsx"
INPUT_TISSUE = "susceptibility score_Midrib.xlsx"
INPUT_QPCR = "qpcr_leaf_midrib_allocation.csv"

SEVERE_THRESHOLD_DEFAULT = 3.0
BOOTSTRAPS_DEFAULT = 10000
RANDOM_SEED_DEFAULT = 20260418


@dataclass
class CanopyModelSummary:
    n_complete_case: int
    r2_lambda_linear: float
    r2_greenhouse_quadratic_full: float
    r2_greenhouse_quadratic_slope_only: float


def require_file(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(f"Missing required input file: {path}")


def safe_number(value):
    if value is None:
        return None
    if isinstance(value, (np.integer, int)):
        return int(value)
    if isinstance(value, (np.floating, float)):
        if math.isnan(float(value)) or math.isinf(float(value)):
            return None
        return float(value)
    return value


def jsd_base2(p: np.ndarray, q: np.ndarray) -> float:
    """Jensen-Shannon divergence using log base 2."""
    p = np.asarray(p, dtype=float)
    q = np.asarray(q, dtype=float)
    if p.ndim != 1 or q.ndim != 1 or p.shape != q.shape:
        raise ValueError("p and q must be one-dimensional arrays with the same shape.")
    if np.any(p < 0) or np.any(q < 0):
        raise ValueError("Probability vectors cannot contain negative values.")
    p_sum = p.sum()
    q_sum = q.sum()
    if p_sum <= 0 or q_sum <= 0:
        return np.nan
    p = p / p_sum
    q = q / q_sum
    m = 0.5 * (p + q)

    def kl(a, b):
        mask = a > 0
        return float(np.sum(a[mask] * (np.log2(a[mask]) - np.log2(b[mask]))))

    return 0.5 * kl(p, m) + 0.5 * kl(q, m)


def odds_ratio_ci(a, b, c, d):
    """Haldane-Anscombe corrected odds ratio and Wald 95% CI."""
    aa, bb, cc, dd = [x + 0.5 for x in [a, b, c, d]]
    estimate = (aa * dd) / (bb * cc)
    se = math.sqrt(1 / aa + 1 / bb + 1 / cc + 1 / dd)
    z = 1.959963984540054
    low = math.exp(math.log(estimate) - z * se)
    high = math.exp(math.log(estimate) + z * se)
    return estimate, low, high


def bootstrap_difference(x, y, statistic="mean", n_boot=BOOTSTRAPS_DEFAULT, seed=RANDOM_SEED_DEFAULT):
    """Independent bootstrap for leaf minus midrib differences."""
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    x = x[~np.isnan(x)]
    y = y[~np.isnan(y)]
    if len(x) == 0 or len(y) == 0:
        return np.nan, np.nan, np.nan, np.nan

    rng = np.random.default_rng(seed)
    x_samples = rng.choice(x, size=(int(n_boot), len(x)), replace=True)
    y_samples = rng.choice(y, size=(int(n_boot), len(y)), replace=True)

    if statistic == "mean":
        observed = float(np.mean(x) - np.mean(y))
        draws = x_samples.mean(axis=1) - y_samples.mean(axis=1)
    elif statistic == "median":
        observed = float(np.median(x) - np.median(y))
        draws = np.median(x_samples, axis=1) - np.median(y_samples, axis=1)
    else:
        raise ValueError("statistic must be 'mean' or 'median'.")

    low, high = np.quantile(draws, [0.025, 0.975])
    return observed, float(low), float(high), float(np.std(draws, ddof=1))


def read_inputs(data_dir: Path):
    architecture_path = data_dir / INPUT_ARCHITECTURE
    tissue_path = data_dir / INPUT_TISSUE
    qpcr_path = data_dir / INPUT_QPCR
    for path in (architecture_path, tissue_path, qpcr_path):
        require_file(path)

    architecture = pd.read_excel(architecture_path)
    tissue = pd.read_excel(tissue_path)
    qpcr = pd.read_csv(qpcr_path)
    return architecture, tissue, qpcr


def build_input_audit(architecture, tissue, qpcr):
    rows = [
        {
            "input": INPUT_ARCHITECTURE,
            "rows": len(architecture),
            "columns": architecture.shape[1],
            "note": "Canopy-angle records with linked cultivar-level greenhouse severity annotations.",
        },
        {
            "input": INPUT_TISSUE,
            "rows": len(tissue),
            "columns": tissue.shape[1],
            "note": "Detached-leaf lesion-score records for leaf blade and midrib tissues.",
        },
        {
            "input": INPUT_QPCR,
            "rows": len(qpcr),
            "columns": qpcr.shape[1],
            "note": "qRT-PCR fold-change summaries used to derive tissue-level defense allocation.",
        },
    ]
    return pd.DataFrame(rows)


def build_canopy_features(df):
    required = {"cultivar", "leaf_no", "angle_deg", "GH_FSP35"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"{INPUT_ARCHITECTURE} missing required columns: {sorted(missing)}")

    records = []
    for cultivar, group in df.groupby("cultivar", dropna=True):
        g = group.copy()
        g["leaf_no"] = pd.to_numeric(g["leaf_no"], errors="coerce")
        g["angle_deg"] = pd.to_numeric(g["angle_deg"], errors="coerce")
        g["GH_FSP35"] = pd.to_numeric(g["GH_FSP35"], errors="coerce")
        g = g.dropna(subset=["leaf_no", "angle_deg"])
        if g.empty:
            continue

        leaf_means = g.groupby("leaf_no")["angle_deg"].mean().sort_index()
        if len(leaf_means) >= 2:
            slope, intercept = np.polyfit(
                leaf_means.index.to_numpy(dtype=float),
                leaf_means.to_numpy(dtype=float),
                1,
            )
        else:
            slope, intercept = np.nan, np.nan

        mean_basal_angle = float(g["angle_deg"].mean())
        theta = np.deg2rad(mean_basal_angle)
        greenhouse_values = g["GH_FSP35"].dropna()

        records.append({
            "cultivar": str(cultivar),
            "mean_profile_angle": float(leaf_means.mean()),
            "profile_sd": float(leaf_means.std(ddof=1)),
            "min_angle": float(leaf_means.min()),
            "max_angle": float(leaf_means.max()),
            "angle_range": float(leaf_means.max() - leaf_means.min()),
            "leaf_slope": float(slope),
            "intercept": float(intercept),
            "mean_basal_angle": mean_basal_angle,
            "runoff_index": float(np.cos(theta)),
            "retention_index": float(np.sin(theta)),
            "GH_FSP35_mean": float(greenhouse_values.mean()) if len(greenhouse_values) else np.nan,
            "n_angle_records": int(len(g)),
            "n_greenhouse_annotation_records": int(len(greenhouse_values)),
        })

    return pd.DataFrame(records).sort_values("cultivar").reset_index(drop=True)


def build_angle_by_leaf(df):
    required = {"cultivar", "leaf_no", "angle_deg"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"{INPUT_ARCHITECTURE} missing required columns: {sorted(missing)}")
    d = df.copy()
    d["leaf_no"] = pd.to_numeric(d["leaf_no"], errors="coerce")
    d["angle_deg"] = pd.to_numeric(d["angle_deg"], errors="coerce")
    d = d.dropna(subset=["cultivar", "leaf_no", "angle_deg"])
    out = (
        d.groupby(["cultivar", "leaf_no"], dropna=True)["angle_deg"]
        .agg(["count", "mean", "std", "min", "max"])
        .reset_index()
        .rename(columns={
            "count": "n",
            "mean": "mean_angle_deg",
            "std": "sd_angle_deg",
            "min": "min_angle_deg",
            "max": "max_angle_deg",
        })
    )
    return out.sort_values(["cultivar", "leaf_no"]).reset_index(drop=True)


def build_pca_tables(angle_by_leaf):
    profile = angle_by_leaf.pivot(index="cultivar", columns="leaf_no", values="mean_angle_deg")
    profile = profile.sort_index(axis=1)
    used_column_means = profile.mean(axis=0)
    profile_filled = profile.fillna(used_column_means)
    n_components = min(2, profile_filled.shape[0], profile_filled.shape[1])
    if n_components < 2:
        return pd.DataFrame(), pd.DataFrame()

    pca = PCA(n_components=2)
    scores = pca.fit_transform(profile_filled.to_numpy(dtype=float))
    score_df = pd.DataFrame({
        "cultivar": profile_filled.index.astype(str),
        "PC1": scores[:, 0],
        "PC2": scores[:, 1],
    })
    explained = pd.DataFrame({
        "component": ["PC1", "PC2"],
        "explained_variance_ratio": pca.explained_variance_ratio_,
    })
    return score_df, explained


def fit_canopy_models(features, eps=1e-6):
    complete = features.dropna(
        subset=["GH_FSP35_mean", "mean_profile_angle", "leaf_slope", "profile_sd", "mean_basal_angle"]
    ).reset_index(drop=True)
    if len(complete) < 3:
        raise ValueError("Not enough complete-case genotypes for canopy models.")

    gh = complete["GH_FSP35_mean"].to_numpy(dtype=float)
    gh_max = float(np.nanmax(gh))
    if gh_max <= 0:
        raise ValueError("Maximum greenhouse severity must be positive.")
    complete["p_severe_GH_scaled"] = gh / gh_max
    complete["lambda_arch"] = -np.log(complete["p_severe_GH_scaled"].to_numpy(dtype=float) + float(eps))

    predictors = ["mean_basal_angle", "leaf_slope", "profile_sd"]
    model_lambda = LinearRegression().fit(complete[predictors], complete["lambda_arch"])
    complete["lambda_arch_pred"] = model_lambda.predict(complete[predictors])
    r2_lambda = r2_score(complete["lambda_arch"], complete["lambda_arch_pred"])

    slope = complete["leaf_slope"].to_numpy(dtype=float)
    full_x = np.column_stack([
        complete["mean_profile_angle"].to_numpy(dtype=float),
        slope,
        complete["profile_sd"].to_numpy(dtype=float),
        slope ** 2,
    ])
    y = complete["GH_FSP35_mean"].to_numpy(dtype=float)
    model_full = LinearRegression().fit(full_x, y)
    complete["GH_pred_quadratic_full"] = model_full.predict(full_x)
    r2_full = r2_score(y, complete["GH_pred_quadratic_full"])

    slope_x = np.column_stack([slope, slope ** 2])
    model_slope = LinearRegression().fit(slope_x, y)
    complete["GH_pred_quadratic_slope_only"] = model_slope.predict(slope_x)
    r2_slope = r2_score(y, complete["GH_pred_quadratic_slope_only"])

    summary = CanopyModelSummary(
        n_complete_case=int(len(complete)),
        r2_lambda_linear=float(r2_lambda),
        r2_greenhouse_quadratic_full=float(r2_full),
        r2_greenhouse_quadratic_slope_only=float(r2_slope),
    )
    model_table = pd.DataFrame([
        {"model": "lambda_arch_linear", "response": "lambda_arch", "n_complete_case": len(complete), "r2": r2_lambda},
        {"model": "greenhouse_quadratic_full", "response": "GH_FSP35_mean", "n_complete_case": len(complete), "r2": r2_full},
        {"model": "greenhouse_quadratic_slope_only", "response": "GH_FSP35_mean", "n_complete_case": len(complete), "r2": r2_slope},
    ])
    return complete, model_table, summary


def standardize_tissue_table(df):
    required = {"cultivar", "tissue", "FSP35_score"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"{INPUT_TISSUE} missing required columns: {sorted(missing)}")

    d = df.copy()
    d["cultivar"] = d["cultivar"].astype(str).str.strip()
    d["tissue"] = d["tissue"].astype(str).str.strip().str.capitalize()
    d["tissue"] = d["tissue"].replace({"Blade": "Leaf", "Lamina": "Leaf"})
    d["FSP35_score"] = pd.to_numeric(d["FSP35_score"], errors="coerce")
    d = d[d["tissue"].isin(["Leaf", "Midrib"])].copy()
    return d


def compute_tissue_probabilities(tissue, severe_threshold=SEVERE_THRESHOLD_DEFAULT):
    d = standardize_tissue_table(tissue)
    d = d.dropna(subset=["cultivar", "tissue"]).copy()

    audit = (
        d.assign(
            score_nonmissing=d["FSP35_score"].notna(),
            severe=d["FSP35_score"].ge(float(severe_threshold)) & d["FSP35_score"].notna(),
        )
        .groupby(["cultivar", "tissue"], dropna=True)
        .agg(
            total_rows=("FSP35_score", "size"),
            nonmissing_score_rows=("score_nonmissing", "sum"),
            missing_score_rows=("score_nonmissing", lambda x: int((~x).sum())),
            severe_score_rows=("severe", "sum"),
        )
        .reset_index()
    )
    audit["nonsevere_score_rows"] = audit["nonmissing_score_rows"] - audit["severe_score_rows"]
    audit["p_severe"] = np.where(
        audit["nonmissing_score_rows"] > 0,
        audit["severe_score_rows"] / audit["nonmissing_score_rows"],
        np.nan,
    )

    wide_p = audit.pivot(index="cultivar", columns="tissue", values="p_severe").rename(
        columns={"Leaf": "p_leaf", "Midrib": "p_midrib"}
    )
    wide_n = audit.pivot(index="cultivar", columns="tissue", values="nonmissing_score_rows").rename(
        columns={"Leaf": "n_leaf", "Midrib": "n_midrib"}
    )
    wide_missing = audit.pivot(index="cultivar", columns="tissue", values="missing_score_rows").rename(
        columns={"Leaf": "missing_leaf_scores", "Midrib": "missing_midrib_scores"}
    )

    out = pd.concat([wide_n, wide_missing, wide_p], axis=1).reset_index()
    out["delta_p_leaf_minus_midrib"] = out["p_leaf"] - out["p_midrib"]
    return out.sort_values("cultivar").reset_index(drop=True), audit.sort_values(["cultivar", "tissue"]).reset_index(drop=True)


def compute_tissue_ordinal_tests(tissue, severe_threshold=SEVERE_THRESHOLD_DEFAULT, n_boot=BOOTSTRAPS_DEFAULT, seed=RANDOM_SEED_DEFAULT):
    d = standardize_tissue_table(tissue)
    d = d.dropna(subset=["cultivar", "tissue", "FSP35_score"]).copy()

    rows = []
    for cultivar, group in d.groupby("cultivar", dropna=True):
        leaf = group.loc[group["tissue"] == "Leaf", "FSP35_score"].dropna().to_numpy(dtype=float)
        midrib = group.loc[group["tissue"] == "Midrib", "FSP35_score"].dropna().to_numpy(dtype=float)
        if len(leaf) == 0 or len(midrib) == 0:
            continue

        mw = mannwhitneyu(leaf, midrib, alternative="two-sided")
        u_stat = float(mw.statistic)
        rank_biserial = (2 * u_stat) / (len(leaf) * len(midrib)) - 1

        mean_diff, mean_low, mean_high, mean_boot_sd = bootstrap_difference(
            leaf, midrib, statistic="mean", n_boot=n_boot, seed=seed
        )
        median_diff, median_low, median_high, median_boot_sd = bootstrap_difference(
            leaf, midrib, statistic="median", n_boot=n_boot, seed=seed + 17
        )

        leaf_severe = int(np.sum(leaf >= severe_threshold))
        leaf_nonsevere = int(np.sum(leaf < severe_threshold))
        midrib_severe = int(np.sum(midrib >= severe_threshold))
        midrib_nonsevere = int(np.sum(midrib < severe_threshold))
        _, fisher_p = fisher_exact(
            [[leaf_severe, leaf_nonsevere], [midrib_severe, midrib_nonsevere]],
            alternative="two-sided",
        )
        odds_ratio, odds_low, odds_high = odds_ratio_ci(leaf_severe, leaf_nonsevere, midrib_severe, midrib_nonsevere)

        rows.append({
            "genotype": str(cultivar),
            "leaf_n": int(len(leaf)),
            "midrib_n": int(len(midrib)),
            "leaf_mean_score": float(np.mean(leaf)),
            "midrib_mean_score": float(np.mean(midrib)),
            "leaf_median_score": float(np.median(leaf)),
            "midrib_median_score": float(np.median(midrib)),
            "mean_diff_leaf_minus_midrib": mean_diff,
            "mean_diff_95ci_low": mean_low,
            "mean_diff_95ci_high": mean_high,
            "median_diff_leaf_minus_midrib": median_diff,
            "median_diff_95ci_low": median_low,
            "median_diff_95ci_high": median_high,
            "mann_whitney_p": float(mw.pvalue),
            "rank_biserial_correlation": float(rank_biserial),
            "p_severe_leaf": float(np.mean(leaf >= severe_threshold)),
            "p_severe_midrib": float(np.mean(midrib >= severe_threshold)),
            "fisher_exact_p": float(fisher_p),
            "odds_ratio_leaf_vs_midrib": float(odds_ratio),
            "odds_ratio_95ci_low": float(odds_low),
            "odds_ratio_95ci_high": float(odds_high),
        })

    out = pd.DataFrame(rows).sort_values("genotype").reset_index(drop=True)
    if len(out):
        out["mann_whitney_fdr_p"] = multipletests(out["mann_whitney_p"], method="fdr_bh")[1]
        out["fisher_exact_fdr_p"] = multipletests(out["fisher_exact_p"], method="fdr_bh")[1]
    return out


def fit_pooled_ordinal_models(tissue):
    d = standardize_tissue_table(tissue)
    d = d.dropna(subset=["cultivar", "tissue", "FSP35_score"]).copy()
    d["score_int"] = d["FSP35_score"].round().astype(int)

    cultivar_dummies = pd.get_dummies(d["cultivar"], prefix="cultivar", drop_first=True, dtype=float)
    tissue_leaf = (d["tissue"] == "Leaf").astype(float).rename("tissue_leaf")

    exog0 = cultivar_dummies.copy()
    exog1 = pd.concat([cultivar_dummies, tissue_leaf], axis=1)
    interaction = cultivar_dummies.mul(tissue_leaf, axis=0)
    interaction.columns = [f"{col}:tissue_leaf" for col in interaction.columns]
    exog2 = pd.concat([cultivar_dummies, tissue_leaf, interaction], axis=1)

    y = d["score_int"]

    res0 = OrderedModel(y, exog0, distr="logit").fit(method="bfgs", disp=False)
    res1 = OrderedModel(y, exog1, distr="logit").fit(method="bfgs", disp=False)
    res2 = OrderedModel(y, exog2, distr="logit").fit(method="bfgs", disp=False)

    lr_tissue = 2 * (res1.llf - res0.llf)
    df_tissue = exog1.shape[1] - exog0.shape[1]
    p_tissue = chi2.sf(lr_tissue, df_tissue)

    lr_interaction = 2 * (res2.llf - res1.llf)
    df_interaction = exog2.shape[1] - exog1.shape[1]
    p_interaction = chi2.sf(lr_interaction, df_interaction)

    model_summary = pd.DataFrame([
        {"model": "cultivar_only", "n_obs": len(d), "log_likelihood": res0.llf, "aic": res0.aic, "n_predictor_terms": exog0.shape[1]},
        {"model": "cultivar_plus_tissue", "n_obs": len(d), "log_likelihood": res1.llf, "aic": res1.aic, "n_predictor_terms": exog1.shape[1]},
        {"model": "cultivar_tissue_interaction", "n_obs": len(d), "log_likelihood": res2.llf, "aic": res2.aic, "n_predictor_terms": exog2.shape[1]},
    ])

    tests = pd.DataFrame([
        {"comparison": "add_tissue_to_cultivar_model", "lr_statistic": lr_tissue, "df": df_tissue, "p_value": p_tissue},
        {"comparison": "add_cultivar_by_tissue_interaction", "lr_statistic": lr_interaction, "df": df_interaction, "p_value": p_interaction},
    ])

    coef = pd.DataFrame({
        "term": res1.params.index,
        "estimate": res1.params.values,
        "model": "cultivar_plus_tissue",
    })
    return model_summary, tests, coef


def compute_defense_allocation(qpcr):
    required = {"cultivar", "leaf_fc", "mid_fc"}
    missing = required - set(qpcr.columns)
    if missing:
        raise ValueError(f"{INPUT_QPCR} missing required columns: {sorted(missing)}")

    d = qpcr.copy()
    d["cultivar"] = d["cultivar"].astype(str).str.strip()
    d["leaf_fc"] = pd.to_numeric(d["leaf_fc"], errors="coerce")
    d["mid_fc"] = pd.to_numeric(d["mid_fc"], errors="coerce")
    d = d.dropna(subset=["cultivar", "leaf_fc", "mid_fc"]).copy()
    denominator = d["leaf_fc"] + d["mid_fc"]
    d = d[denominator > 0].copy()
    denominator = d["leaf_fc"] + d["mid_fc"]
    d["p_defense_leaf"] = d["leaf_fc"] / denominator
    d["p_defense_midrib"] = 1.0 - d["p_defense_leaf"]

    allocation = (
        d.groupby("cultivar", dropna=True)[["p_defense_leaf", "p_defense_midrib"]]
        .mean()
        .reset_index()
        .sort_values("cultivar")
    )
    allocation["cultivar"] = allocation["cultivar"].astype(str)
    return allocation, d


def build_alignment_table(tissue_probabilities, defense_allocation):
    merged = tissue_probabilities.merge(defense_allocation, on="cultivar", how="inner")
    merged = merged.dropna(subset=["p_leaf", "p_midrib", "p_defense_leaf", "p_defense_midrib"]).copy()

    total_attack = merged["p_leaf"] + merged["p_midrib"]
    merged = merged[total_attack > 0].copy()
    total_attack = merged["p_leaf"] + merged["p_midrib"]

    merged["p_attack_leaf_normalized"] = merged["p_leaf"] / total_attack
    merged["p_attack_midrib_normalized"] = merged["p_midrib"] / total_attack

    jsd_values = []
    for _, row in merged.iterrows():
        attack = np.array([row["p_attack_leaf_normalized"], row["p_attack_midrib_normalized"]], dtype=float)
        defense = np.array([row["p_defense_leaf"], row["p_defense_midrib"]], dtype=float)
        jsd_values.append(jsd_base2(attack, defense))
    merged["jsd_attack_defense"] = jsd_values
    merged["overall_protection"] = 1.0 - (merged["p_leaf"] + merged["p_midrib"]) / 2.0

    return merged[[
        "cultivar",
        "p_leaf",
        "p_midrib",
        "p_attack_leaf_normalized",
        "p_attack_midrib_normalized",
        "p_defense_leaf",
        "p_defense_midrib",
        "jsd_attack_defense",
        "overall_protection",
        "delta_p_leaf_minus_midrib",
    ]].sort_values("cultivar").reset_index(drop=True)


def build_manuscript_table2(alignment):
    out = alignment.rename(columns={
        "cultivar": "Cultivar",
        "p_leaf": "p(severe) leaf (FSP35)",
        "p_midrib": "p(severe) midrib (FSP35)",
        "p_attack_leaf_normalized": "p(attack) leaf (normalized)",
        "p_attack_midrib_normalized": "p(attack) midrib (normalized)",
        "p_defense_leaf": "p(defense) leaf (FSP53 qPCR)",
        "p_defense_midrib": "p(defense) midrib (FSP53 qPCR)",
        "jsd_attack_defense": "JSD (attack-defense)",
        "overall_protection": "Return (overall protection)",
    })
    return out[[
        "Cultivar",
        "p(severe) leaf (FSP35)",
        "p(severe) midrib (FSP35)",
        "p(attack) leaf (normalized)",
        "p(attack) midrib (normalized)",
        "p(defense) leaf (FSP53 qPCR)",
        "p(defense) midrib (FSP53 qPCR)",
        "JSD (attack-defense)",
        "Return (overall protection)",
    ]]


def build_manuscript_table_s1(tissue_tests):
    cols = [
        "genotype",
        "leaf_n",
        "midrib_n",
        "leaf_mean_score",
        "midrib_mean_score",
        "mean_diff_leaf_minus_midrib",
        "mean_diff_95ci_low",
        "mean_diff_95ci_high",
        "mann_whitney_p",
        "mann_whitney_fdr_p",
        "p_severe_leaf",
        "p_severe_midrib",
        "fisher_exact_p",
    ]
    out = tissue_tests[cols].copy()
    out = out.rename(columns={
        "genotype": "Genotype",
        "leaf_n": "Leaf n",
        "midrib_n": "Midrib n",
        "leaf_mean_score": "Leaf mean score",
        "midrib_mean_score": "Midrib mean score",
        "mean_diff_leaf_minus_midrib": "Mean diff (Leaf-Midrib)",
        "mean_diff_95ci_low": "95% CI low",
        "mean_diff_95ci_high": "95% CI high",
        "mann_whitney_p": "Mann-Whitney p",
        "mann_whitney_fdr_p": "Mann-Whitney FDR p",
        "p_severe_leaf": "p(severe) leaf",
        "p_severe_midrib": "p(severe) midrib",
        "fisher_exact_p": "Fisher exact p",
    })
    return out


def format_sheet(sheet):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for col_idx, column_cells in enumerate(sheet.columns, start=1):
        max_len = 10
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, min(len(str(value)), 50))
            if isinstance(value, float):
                cell.number_format = "0.000000"
        sheet.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 32))


def add_dataframe_sheet(workbook, name, dataframe):
    sheet = workbook.create_sheet(name)
    for col_idx, col in enumerate(dataframe.columns, start=1):
        sheet.cell(row=1, column=col_idx, value=str(col))
    for row_idx, (_, row) in enumerate(dataframe.iterrows(), start=2):
        for col_idx, col in enumerate(dataframe.columns, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=safe_number(row[col]))
    format_sheet(sheet)


def write_workbook(
    output_path,
    readme_rows,
    input_audit,
    angle_by_leaf,
    canopy_features,
    pca_scores,
    pca_explained,
    canopy_complete,
    canopy_model_table,
    tissue_probability_audit,
    tissue_probabilities,
    tissue_tests,
    ordinal_model_summary,
    ordinal_lr_tests,
    ordinal_coefficients,
    qpcr_allocation_records,
    defense_allocation,
    alignment,
    manuscript_table2,
    manuscript_table_s1,
):
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    readme = workbook.create_sheet("00_README")
    readme["A1"] = "Supplementary data workbook"
    readme["A1"].font = Font(bold=True, size=14)
    readme["A3"] = "Notes"
    readme["A3"].font = Font(bold=True)
    for idx, text in enumerate(readme_rows, start=4):
        readme.cell(row=idx, column=1, value=text)
        readme.cell(row=idx, column=1).alignment = Alignment(wrap_text=True)
    readme.column_dimensions["A"].width = 120

    add_dataframe_sheet(workbook, "01_Input_Audit", input_audit)
    add_dataframe_sheet(workbook, "02_Angle_by_Leaf", angle_by_leaf)
    add_dataframe_sheet(workbook, "03_Canopy_Traits", canopy_features)
    add_dataframe_sheet(workbook, "04_Canopy_PCA_Scores", pca_scores)
    add_dataframe_sheet(workbook, "05_Canopy_PCA_Variance", pca_explained)
    add_dataframe_sheet(workbook, "06_Canopy_Model_Data", canopy_complete)
    add_dataframe_sheet(workbook, "07_Canopy_Model_Summary", canopy_model_table)
    add_dataframe_sheet(workbook, "08_Tissue_Probability_Audit", tissue_probability_audit)
    add_dataframe_sheet(workbook, "09_Tissue_Probabilities", tissue_probabilities)
    add_dataframe_sheet(workbook, "10_Tissue_Ordinal_Tests", tissue_tests)
    add_dataframe_sheet(workbook, "11_Pooled_Ordinal_Models", ordinal_model_summary)
    add_dataframe_sheet(workbook, "12_Pooled_Ordinal_Tests", ordinal_lr_tests)
    add_dataframe_sheet(workbook, "13_Pooled_Ordinal_Coefficients", ordinal_coefficients)
    add_dataframe_sheet(workbook, "14_qPCR_Allocation_Records", qpcr_allocation_records)
    add_dataframe_sheet(workbook, "15_Defense_Allocation", defense_allocation)
    add_dataframe_sheet(workbook, "16_Attack_Defense_JSD", alignment)
    add_dataframe_sheet(workbook, "17_Manuscript_Table2", manuscript_table2)
    add_dataframe_sheet(workbook, "18_Manuscript_TableS1", manuscript_table_s1)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def run_analysis(data_dir, out_dir, severe_threshold, n_boot, seed, overwrite):
    architecture, tissue, qpcr = read_inputs(data_dir)

    input_audit = build_input_audit(architecture, tissue, qpcr)
    angle_by_leaf = build_angle_by_leaf(architecture)
    canopy_features = build_canopy_features(architecture)
    pca_scores, pca_explained = build_pca_tables(angle_by_leaf)
    canopy_complete, canopy_model_table, canopy_summary = fit_canopy_models(canopy_features)

    tissue_probabilities, tissue_probability_audit = compute_tissue_probabilities(
        tissue, severe_threshold=severe_threshold
    )
    tissue_tests = compute_tissue_ordinal_tests(
        tissue, severe_threshold=severe_threshold, n_boot=n_boot, seed=seed
    )
    ordinal_model_summary, ordinal_lr_tests, ordinal_coefficients = fit_pooled_ordinal_models(tissue)

    defense_allocation, qpcr_allocation_records = compute_defense_allocation(qpcr)
    alignment = build_alignment_table(tissue_probabilities, defense_allocation)
    manuscript_table2 = build_manuscript_table2(alignment)
    manuscript_table_s1 = build_manuscript_table_s1(tissue_tests)

    output_path = out_dir / "Supplementary_Data_S1.xlsx"
    if output_path.exists() and not overwrite:
        raise FileExistsError(f"{output_path} already exists. Use --overwrite to replace it.")

    readme_rows = [
        "This workbook was generated from the three input data files using run_pipeline.py.",
        "Rows with missing FSP35 lesion scores are excluded from denominators when calculating severe-infection probabilities.",
        f"Severe infection threshold: FSP35 score >= {severe_threshold}.",
        f"Bootstrap iterations for tissue effect-size confidence intervals: {n_boot}.",
        f"Canopy complete-case n: {canopy_summary.n_complete_case}.",
        f"Canopy model R2 values: lambda linear = {canopy_summary.r2_lambda_linear:.3f}; greenhouse quadratic full = {canopy_summary.r2_greenhouse_quadratic_full:.3f}; greenhouse quadratic slope-only = {canopy_summary.r2_greenhouse_quadratic_slope_only:.3f}.",
    ]

    write_workbook(
        output_path=output_path,
        readme_rows=readme_rows,
        input_audit=input_audit,
        angle_by_leaf=angle_by_leaf,
        canopy_features=canopy_features,
        pca_scores=pca_scores,
        pca_explained=pca_explained,
        canopy_complete=canopy_complete,
        canopy_model_table=canopy_model_table,
        tissue_probability_audit=tissue_probability_audit,
        tissue_probabilities=tissue_probabilities,
        tissue_tests=tissue_tests,
        ordinal_model_summary=ordinal_model_summary,
        ordinal_lr_tests=ordinal_lr_tests,
        ordinal_coefficients=ordinal_coefficients,
        qpcr_allocation_records=qpcr_allocation_records,
        defense_allocation=defense_allocation,
        alignment=alignment,
        manuscript_table2=manuscript_table2,
        manuscript_table_s1=manuscript_table_s1,
    )

    return output_path, canopy_summary, ordinal_lr_tests, manuscript_table2


def main():
    parser = argparse.ArgumentParser(description="Generate the supplementary data workbook.")
    parser.add_argument("--data-dir", default=".", help="Directory containing the three input files.")
    parser.add_argument("--out-dir", default="outputs", help="Directory for the output workbook.")
    parser.add_argument("--severe-threshold", type=float, default=SEVERE_THRESHOLD_DEFAULT, help="Threshold for severe infection.")
    parser.add_argument("--bootstrap-iterations", type=int, default=BOOTSTRAPS_DEFAULT, help="Number of bootstrap iterations.")
    parser.add_argument("--seed", type=int, default=RANDOM_SEED_DEFAULT, help="Random seed for bootstrap sampling.")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing output workbook.")
    args = parser.parse_args()

    data_dir = Path(args.data_dir).expanduser().resolve()
    out_dir = Path(args.out_dir).expanduser().resolve()

    output_path, canopy_summary, ordinal_tests, table2 = run_analysis(
        data_dir=data_dir,
        out_dir=out_dir,
        severe_threshold=args.severe_threshold,
        n_boot=args.bootstrap_iterations,
        seed=args.seed,
        overwrite=args.overwrite,
    )

    print(f"Wrote {output_path}")
    print(f"Canopy complete-case n = {canopy_summary.n_complete_case}")
    print(f"Canopy R2, quadratic full model = {canopy_summary.r2_greenhouse_quadratic_full:.3f}")
    print(f"Canopy R2, quadratic slope-only model = {canopy_summary.r2_greenhouse_quadratic_slope_only:.3f}")
    if len(ordinal_tests):
        for _, row in ordinal_tests.iterrows():
            print(f"{row['comparison']}: p = {row['p_value']:.6g}")
    if len(table2):
        print("Corrected Table 2 values were written to sheet 17_Manuscript_Table2.")


if __name__ == "__main__":
    main()
